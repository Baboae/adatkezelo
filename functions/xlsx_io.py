#functions/xlsx.io.py:

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
from dataclasses import asdict, is_dataclass
from typing import List, Type, TypeVar, Dict, Any
import json

T = TypeVar("T")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col[0].column_letter].width = max_len + 2


def _style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="4F81BD")  # kék
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill


def _fill_row(ws, row_idx: int, fill: PatternFill) -> None:
    for cell in ws[row_idx]:
        cell.fill = fill


def _sheet_from_rows(wb: Workbook, name: str, rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title=name)
    if not rows:
        return

    # fejléc
    headers = list(rows[0].keys())
    ws.append(headers)
    _style_header(ws)

    # adatsorok + alap háttérszín
    default_fill = PatternFill("solid", fgColor="DDDDDD")
    for r in rows:
        ws.append([r[h] for h in headers])
        _fill_row(ws, ws.max_row, default_fill)

    _auto_width(ws)


def save_xlsx(sheets: dict, filename: str) -> None:
    """
    sheets: {"SheetName": [items], ...}
    - items lehet dataclass vagy dict; dataclass -> asdict
    - beágyazott dict/list -> JSON string
    - alap háttérszín minden cellának
    - oszlopszélesség automatikus
    - Participants lapon: rating_change/reputation_change zöld/piros, finish_position zöld/piros a start_position-hoz képest
    """
    wb = Workbook()
    wb.remove(wb.active)

    # előkészítés: dataclass -> dict, nested -> json string
    prepared: Dict[str, List[Dict[str, Any]]] = {}
    for name, items in sheets.items():
        rows: List[Dict[str, Any]] = []
        for it in items:
            row = asdict(it) if is_dataclass(it) else (dict(it) if isinstance(it, dict) else None)
            if row is None:
                # nem dict és nem dataclass: tegyük be egyetlen "value" kulcs alatt
                row = {"value": it}
            # lapítás: beágyazott struktúrák JSON stringgé
            flat = {}
            for k, v in row.items():
                if isinstance(v, (dict, list)):
                    flat[k] = json.dumps(v, ensure_ascii=False)
                else:
                    flat[k] = v
            rows.append(flat)
        prepared[name] = rows

    # írjuk a lapokat
    for name, rows in prepared.items():
        _sheet_from_rows(wb, name, rows)

    # speciális színezés a Participants lapon
    if "Participants" in wb.sheetnames:
        ws = wb["Participants"]
        if ws.max_row >= 2:
            # header -> index mapping
            headers = [c.value for c in ws[1]]
            try:
                rating_change_idx = headers.index("rating_change") + 1
            except ValueError:
                rating_change_idx = None
            try:
                reputation_change_idx = headers.index("reputation_change") + 1
            except ValueError:
                reputation_change_idx = None
            try:
                start_idx = headers.index("start_position") + 1
            except ValueError:
                start_idx = None
            try:
                finish_idx = headers.index("finish_position") + 1
            except ValueError:
                finish_idx = None

            green_fill = PatternFill("solid", fgColor="C6EFCE")
            red_fill = PatternFill("solid", fgColor="FFC7CE")

            # színezés: rating_change és reputation_change
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if rating_change_idx is not None:
                    cell = row[rating_change_idx - 1]
                    try:
                        val = float(cell.value)
                        if val > 0:
                            cell.fill = green_fill
                        elif val < 0:
                            cell.fill = red_fill
                    except Exception:
                        pass
                if reputation_change_idx is not None:
                    cell = row[reputation_change_idx - 1]
                    try:
                        val = float(cell.value)
                        if val > 0:
                            cell.fill = green_fill
                        elif val < 0:
                            cell.fill = red_fill
                    except Exception:
                        pass

                # pozíció változás: finish vs start
                if start_idx is not None and finish_idx is not None:
                    start_val = row[start_idx - 1].value
                    finish_val = row[finish_idx - 1].value
                    if isinstance(start_val, int) and isinstance(finish_val, int):
                        diff = finish_val - start_val
                        if diff < 0:
                            row[finish_idx - 1].fill = green_fill
                        elif diff > 0:
                            row[finish_idx - 1].fill = red_fill

            _auto_width(ws)  # szélesség újraszámolás a színezés után is

    # Laps lapon: semleges formázás már megvan; opcionálisan pozíció színezés elhagyva, mert nincs összehasonlítási alap

    path = Path("created/xlsxs") / filename
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def load_xlsx(filename: str, sheet: str, cls: Type[T]) -> List[T]:
    path = Path("created/xlsxs") / filename
    wb = load_workbook(path)
    ws = wb[sheet]
    rows = list(ws.rows)
    if not rows:
        return []
    headers = [cell.value for cell in rows[0]]
    result = []
    for row in rows[1:]:
        values = [cell.value for cell in row]
        data = dict(zip(headers, values))
        result.append(cls(**data))
    return result
